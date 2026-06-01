from dataclasses import dataclass, field
from pathlib import Path
import pdfplumber
import openpyxl

@dataclass
class ParsedDocument:
    text: str
    tables: list[list[list[str]]]
    file_type: str

def parse_document(file_path: str) -> ParsedDocument:
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".pdf":
        return _parse_pdf(path)
    elif suffix in (".xlsx", ".xls"):
        return _parse_excel(path)
    raise ValueError(f"Unsupported file type: {suffix}. 지원 형식: .pdf, .xlsx, .xls")

def _parse_pdf(path: Path) -> ParsedDocument:
    pages_text: list[str] = []
    tables: list[list[list[str]]] = []

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                pages_text.append(text)
            for table in (page.extract_tables() or []):
                normalized = [
                    [str(cell or "").strip() for cell in row]
                    for row in table
                ]
                tables.append(normalized)

    return ParsedDocument(text="\n".join(pages_text), tables=tables, file_type="pdf")

def _parse_excel(path: Path) -> ParsedDocument:
    wb = openpyxl.load_workbook(path, data_only=True)
    all_tables: list[list[list[str]]] = []
    text_lines: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        for row in ws.iter_rows():
            cells = [str(cell.value or "").strip() for cell in row]
            if any(c for c in cells):
                rows.append(cells)
                text_lines.append("\t".join(cells))
        if rows:
            all_tables.append(rows)

    return ParsedDocument(text="\n".join(text_lines), tables=all_tables, file_type="excel")
